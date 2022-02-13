import openpyxl
import sys
from unipath import Path
from openpyxl.utils import get_column_letter

VERBOSE = False


def del_data_except_col_header(ws_obj, header_cols=1):
    """Deletes all the rows, header col by default is 1, so it will skip the first row"""
    amount = ws_obj.max_row - header_cols
    ws_obj.delete_rows(2, amount)


def open_xls(xls_file_name):
    """Returns the WorkBook of specified Name"""
    if not xls_file_name.exists():
        print("The following file does not exists:", xls_file_name)
        print("Please ensure the file exists or the correct filename was entered when "
              "utilizing the '-i | --input_file' option.")
        sys.exit()
    if VERBOSE:
        print("Opening Excel sheet:", xls_file_name)
    return openpyxl.load_workbook(xls_file_name, data_only=True)


def rw_cell(sheet_obj, row, column, value=None):
    """
    Either writes or reads to/from a cell.
    To Erase a cell please use clear_cell
    """
    if isinstance(column, str):
        c_index = column.upper() + str(row)
        cell_obj = sheet_obj[c_index]
        if value is not None:
            cell_obj.value = value
    else:
        cell_obj = sheet_obj.cell(row=row, column=column, value=value)
    return cell_obj.value


def cell_iter_to_list(cell_iter, ignore_empty_cell=True):
    t_list = []
    for cell in cell_iter:
        if not ignore_empty_cell:
            t_list.append(cell.value)
        elif cell.value:
            t_list.append(cell.value)
    return t_list


def clear_cell(sheet_obj, row, column):
    """
    Delete the contents of a Cell
    :param sheet_obj: sheet object where cell is to be deleted
    :param row: row of cell
    :param column: column of cell
    :return:
    """
    sheet_obj.cell(row=row, column=column).value = None


def add_xls_tag(file_name):
    """
    Check the file_name to ensure it has ".xlsx" extension, if not add it
    """
    file_name = str(file_name)
    if file_name[:-5] != ".xlsx":
        return file_name + ".xlsx"
    else:
        return file_name


def next_available_row(sheet_obj, col='A'):
    """
    Returns the number of the next available Row, it determines
    avaibaility based on whether there is an entry the column 'col'
    """
    column = sheet_obj[col]
    for index, cell in enumerate(column):
        if cell.value is None:
            return index + 1
    return len(column) + 1


def save_xls(wb_obj, file_name, out_dir_path=""):
    """Saves the WorkBook to provided Directory and File Name.
    If no filename will save as input file name with _new at end.
    If no out_dir_path, it will save as ISE_downlink_automation_New.xlsx with a random string
    """
    out_dir_path = Path(out_dir_path)
    file_save_string = out_dir_path.child(file_name)
    print("Saving the file to:", file_save_string)
    wb_obj.save(file_save_string)


def save_xls_retry_if_open(wb_obj, file_name, out_dir_path=""):
    out_dir_path = Path(out_dir_path)
    file_open = True
    file_name = add_xls_tag(file_name)
    while file_open:
        try:
            save_xls(wb_obj, file_name, out_dir_path)
            file_open = False
        except PermissionError:
            msg = "ERROR:\n\tThe file %s is open. Please close the file and try again. Hit Enter when ready."
            msg += "\n\tType ignore if you wish to continue without saving.\n"
            response = input(msg % file_name)
            if "ignore" in response:
                return


def add_net_dev_comment_msgs_to_wb(net_dev, wb_obj):
    """
    Adds all the Registered Errors and Comments to the WorkSheet.
    """
    ws_obj = wb_obj["Comments"]
    for i, msg in enumerate(net_dev.cmnt_msgs):
        comment = [net_dev.host, net_dev.hostname, msg[2], net_dev.main_col, msg[1], msg[0]]
        list_of_list_to_ws(ws_obj, [comment])


def list_of_list_to_ws(sheet_obj, list_of_list, fill_down=None):
    """Writes a List of Lists into a ws_obj, can take in a fill_down value and it will add the fill down to the begining of each row"""
    if fill_down is None:
        fill_down = list()
    row = next_available_row(sheet_obj)
    if isinstance(fill_down, str):
        fill_down = [fill_down]
    if isinstance(list_of_list, list):
        for i, list_obj in enumerate(list_of_list):
            col = 1
            for j, val in enumerate(fill_down):
                rw_cell(sheet_obj, row + i, col, val)
                col += 1
            for j, val in enumerate(list_obj):
                if isinstance(val, list):
                    rw_cell(sheet_obj, row + i, col + j, ','.join(val))
                else:
                    rw_cell(sheet_obj, row + i, col + j, str(val))


def list_of_dict_to_ws(sheet_obj, list_of_dicts, fill_down=None):
    """
    Writes a List of Dict into a ws_obj, can take in a fill_down value and it will add the
    fill down to the begining of each row
    """
    row = next_available_row(sheet_obj)
    if isinstance(fill_down, str):
        fill_down = [fill_down]
    elif fill_down is None:
        fill_down = list()
    if isinstance(list_of_dicts, list):
        for i, dict_obj in enumerate(list_of_dicts):
            list_to_write = fill_down+list(dict_obj.values())
            for col, val in enumerate(list_to_write, 1):
                rw_cell(sheet_obj, row + i, col, val)
    else:
        print("check the list_of_dicts_to_ws function")
        print(type(list_of_dicts))
        print(isinstance(list_of_dicts, list))
        print(list_of_dicts)


def set_width(sheet_obj, width, col):
    if isinstance(col, int):
        col = get_column_letter(col)
    sheet_obj.column_dimensions[col].width = width