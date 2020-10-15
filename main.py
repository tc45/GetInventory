"""###v2.0###"""
import socket
from threading import Thread
import traceback
import os
import os.path
import json
import re
from datetime import datetime
import time
import getopt
import sys
from pathlib import Path
import logging
import platform
import netmiko
import openpyxl
import optparse
import textfsm

INPUT_FILE_NAME = "GetInventory - Default.xlsx"

"""###VERBOSE Output###"""
VERBOSE = False

RAW_CLI_OUTPUT = False

TESTING = False

"""###Global Variables###"""
GLBL_KEY_MAP = {}



def main():
    """
    Main Functino to run everything.
    """
    global GLBL_KEY_MAP
    global INPUT_FILE_NAME
    global VERBOSE
    global RAW_CLI_OUTPUT
    update_ntc_templ_path()
    cli_arg = cli_args()
    # change the Input File from teh Default based on CLI Commands
    if cli_arg["input_file"]:
        INPUT_FILE_NAME = cli_arg["input_file"]
    VERBOSE = cli_arg["verbose"]
    ###Print Starting
    spacer = "\n\n" + gen_spacer("#", 1)
    print(spacer + "\t\tStarting GetInventory Script" + spacer)
    # Get all global variables
    print("(1) Getting JSON Data from File")
    GLBL_KEY_MAP = get_json_data_from_file("cmd_xls_key_map.json")
    print("(2) Opening XLS File")
    work_book = open_xls(INPUT_FILE_NAME)
    print("(3) Getting Setup variables")
    setup_vars = get_setup_vars(work_book, cli_arg)
    # Create path if raw_cli_output option was selected
    if cli_arg["raw_cli_output"]:
        RAW_CLI_OUTPUT = cli_arg["raw_cli_output"]
        verify_path(setup_vars["global"]["output_dir"]+"raw_cli")
    print("(4) Reading Network Devices")
    network_devices = read_network_devices(work_book, setup_vars["global"])

    print('(5) Connecting to Devices and capturing commands')
    if VERBOSE:
        print("Excel Row | Host            | Message")
        print(60*"-")
    ###Connects to net_devices
    """If you need to add more functions or to run more commands add them to the function below"""
    if TESTING:
        testing_connection(network_devices, work_book, GLBL_KEY_MAP)
        sys.exit()
    else:
        connect_devices(network_devices, setup_vars)

    print("(6) Saving all the Device Data")
    # Clean up Passwords before that
    remove_passwords(work_book)
    # Saves everything
    """All Save features should be handled by this"""
    save_device_data(network_devices, work_book, setup_vars, GLBL_KEY_MAP)
    print(spacer + "(7) DONE with the script" + spacer)


"""######## Gather Functions that run based on Settings Tab ########"""


## XR Ready
def gather_version(connection, net_dev, count):
    """
    Captures the show version command and saves the textfsm outcome to
    the NetworkDevice. No additional parsing needed at the moment.
    """
    command = "show version"
    log_cmd_textfsm(connection, net_dev, command, count)
    net_dev.read_vers_info()
    show_proc_cpu(connection, net_dev, count)


## XR Ready, Need to test more
def gather_arp(connection, net_dev, count):
    """
    Captures arp information and utilizing the vrf data it parses the
    output to prepare it for extraction to WB.
    """
    vrf_list = get_vrf_names(net_dev, connection, count)
    arp_list = []
    for vrf in vrf_list:
        vrf_string, output = "", ""
        if vrf != "global":
            vrf_string = " vrf " + vrf
        if net_dev.parse_method == "cisco_xr":
            command = "show arp" + vrf_string
            txt_tmpl = "ntc-templates/test_tmpl/cisco_xr_show_arp.textfsm"
            txt_tmpl = mod_dir_based_on_os(txt_tmpl)
        else:
            command = "show ip arp" + vrf_string
            txt_tmpl = None
        output = log_cmd_textfsm(connection, net_dev, command, count, txt_tmpl)
        if isinstance(output, list):
            for arp in output:
                arp["vrf"] = vrf
                if net_dev.parse_method in ["cisco_nxos"]:
                    arp["type"] = "ARPA"
                arp_list.append(arp)
        else:
            arp = {}
            arp['vrf'] = vrf
            arp['address'] = "No ARP Data Found"
            arp_list.append(arp)
    net_dev.show_for_xls["gather_arp"] = arp_list


## XR Ready, Need to test more
def gather_mac(connection, net_dev, count):
    """
    All the parsing of the
    """
    command = "show mac address-table"
    if net_dev.parse_method == "cisco_xr":
        output = []
        cmd = "show l2vpn forwarding bridge-domain {} mac-address location {}"
        locations = get_xr_locations(net_dev, connection, count)
        bg_grp_dmns = get_xr_bg_grp_dmns(net_dev, connection, count)
        txtfsm = "ntc-templates/test_tmpl/cisco_xr_show_l2vpn_bridge-domain_mac.textfsm"
        for dmn in bg_grp_dmns:
            for lctn in locations:
                command = cmd.format(dmn, lctn)
                output += log_cmd_textfsm(connection, net_dev, command, count, txtfsm)
        if not output:
            output += [{"MAC": "No MAC Data"}]
    else:
        output = log_cmd_textfsm(connection, net_dev, command, count)
    if isinstance(output, str) and net_dev.parse_method=='cisco_ios':
        command = "show mac-address-table"
        output = log_cmd_textfsm(connection, net_dev, command, count)
    if isinstance(output, list):
        net_dev.show_for_xls["gather_mac"] = output.copy()
    elif isinstance(output, str):
        output = {}
        output["type"] = "No MAC Address Table results found for device"
        net_dev.show_for_xls["gather_mac"] = [output]


## XR Ready, Need to test more
def gather_interface(connection, net_dev, count):
    """
    Gather Interface information, the function was modified from the
    original script.
    """
    dev_type = net_dev.parse_method
    ####This is from old
    command = "show interface"
    output = net_dev.show_output_json[command].copy()
    log_cmd_textfsm(connection, net_dev, command, count)

    # 'show interface status' capture if necessary
    if dev_type == 'cisco_ios':
        # Only cisco_ios provides useful information here
        command2 = "show interface status"
        log_cmd_textfsm(connection, net_dev, command2, count)
        output2 = net_dev.show_output_json[command2]
    else:
        # all others do not provide the allowed vlans
        output = "NO DATA"

    vrf_info = get_vrf_interfaces_dict(net_dev, connection, count)

    trunks = {}
    switchport_data_found = False
    if isinstance(output2, list):
        switchport_data_found = True
        # If device is a switch, get trunk info into dictionary.
        trunks = get_trunk_dict(net_dev, connection)

    # Write Interface data to spreadsheet 'Interfaces' tab
    if isinstance(output, list):
        for i in output:
            short_if_name = get_short_if_name(i['interface'], dev_type)
            i["short_if"] = short_if_name
            if i['ip_address'] != "":
                i['l2_l3'] = "Layer 3"
                i['trunk_access'] = "Routed"
                if isinstance(vrf_info, list):
                    vrf_name = ""
                    for vrf in vrf_info:
                        for intf in vrf["interfaces"]:
                            t_intf = get_short_if_name(intf, dev_type).lower()
                            if short_if_name.lower() == t_intf:
                                vrf_name = vrf["name"]
                                i['vrf'] = vrf_name
                    if vrf_name == "" and dev_type != "cisco_xr":
                        i['vrf'] = "default"
                    elif vrf_name == "":
                        i['vrf'] = "global"
                else:
                    i['vrf'] = "global"
            if switchport_data_found is True:
                for x in output2:
                    if short_if_name.lower() == x['port'].lower():
                        if x['vlan'].isnumeric():
                            i["vlan"] = x['vlan']
                            i["vlan"] = x['vlan']
                            i["trunk_access"] = "Access"
                            i["l2_l3"] = "Layer 2"
                        elif x['vlan'] == "trunk":
                            i["l2_l3"] = "Layer 2"
                            i["trunk_access"] = "Trunk"
                            # Parse Trunk Details
                            native = get_trunk_details(short_if_name,
                                                       trunks,
                                                       "vlans_native",
                                                       net_dev)
                            allowed = get_trunk_details(short_if_name,
                                                        trunks,
                                                        "vlans_allowed",
                                                        net_dev)
                            not_pruned = get_trunk_details(short_if_name,
                                                           trunks,
                                                           "vlans_not_pruned",
                                                           net_dev)
                            i["native"] = native
                            i["allowed"] = allowed
                            i["not_pruned"] = not_pruned

                        elif x['vlan'] == "routed":
                            i["trunk_access"] = "Routed"
        net_dev.show_for_xls["gather_interface"] = output
        net_dev.interface_count = count_interfaces(net_dev.show_output_json["show interface"])
    elif isinstance(output, str):
        output = {}
        output["status"] = "No Interface Data, have Developer check the script"
        net_dev.show_for_xls["gather_interface"] = [output]


## XR Ready, need to test
def gather_cdp(connection, net_dev, count):
    command = "show cdp neighbor detail"
    output = log_cmd_textfsm(connection, net_dev, command, count)

    #Check to make sure the cdp neigh count matches the detailed count
    if net_dev.parse_method in ['cisco_nxos', 'cisco_ios']:
        command2 = "show cdp neighbor"
        output2 = log_cmd_textfsm(connection, net_dev, command2, count)
        if len(output) != len(output2):
            if VERBOSE:
                print_net_dev_msg(net_dev, "The output of the length of show cdp neigh and thelength of show cdp neighbor details is not the same, please manually gather raw command of both commands")
            net_dev.add_error_msg("The output of the length of show cdp neigh and the length of show cdp neighbor details is not the same, please manually gather raw command of both commands")

    if isinstance(output, list):
        net_dev.show_for_xls["gather_cdp"] = output.copy()
    elif isinstance(output, str):
        output = {}
        output["local_port"] = "No CDP Data"
        net_dev.show_for_xls["gather_cdp"] = [output]


## XR Ready, need to test
def gather_lldp(connection, net_dev, count):
    command = "show lldp neighbor detail"
    txt_tmpl = None
    if net_dev.parse_method == "cisco_xr":
        txt_tmpl = "ntc-templates/test_tmpl/cisco_xr_show_lldp_neighbors_detail.textfsm"
        txt_tmpl = mod_dir_based_on_os(txt_tmpl)
    if net_dev.parse_method == "extreme_exos":
        txt_tmpl = (r"ntc-templates\test_tmpl\extreme_exos_show_lldp_neighbors_detail.textfsm")
    output = log_cmd_textfsm(connection, net_dev, command, count, txt_tmpl)
    if isinstance(output, list):
        net_dev.show_for_xls["gather_lldp"] = output.copy()
    elif isinstance(output, str):
        output = {}
        output["chassis_id"] = "No LLDP Data"
        net_dev.show_for_xls["gather_lldp"] = [output]


## XR Ready, need to test
def gather_route(connection, net_dev, count):
    vrf_list = get_vrf_names(net_dev, connection, count)
    route_list = []
    route_table_present = False
    for vrf in vrf_list:
        vrf_string = ""
        if vrf != "global":
            vrf_string = " vrf " + vrf
        command = "show route" + vrf_string
        output = log_cmd_textfsm(connection, net_dev, command, count)
        if isinstance(output, list):
            for route in output:
                route["cidr"] = route['network'] + "/" + route['mask']
                if "vrf" not in list(route.keys()):
                    route["vrf"] = vrf
                route_list.append(route)
                route_table_present = True
        if not route_table_present:
            route = {}
            default_gateway = re.search(r"\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}", output)
            route["vrf"] = vrf
            route["protocol"] = "Layer 2 only"
            if default_gateway:
                route["nexthop_ip"] = default_gateway[0]
            else:
                output2 = connection.send_command("show run | incl default-gateway")
                default_gateway = re.search(r"\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}", output2)
                if default_gateway:
                    route["nexthop_ip"] = default_gateway[0]
            route_list.append(route)
    net_dev.show_for_xls["gather_route"] = route_list


## XR Ready, Need to test more
def gather_bgp(connection, net_dev, count):
    command = "show ip bgp"
    if net_dev.parse_method == "cisco_xr":
        vrf_names = get_vrf_names(net_dev, connection, count)
        output = []
        txt_fsm = "ntc-templates/test_tmpl/cisco_xr_show_bgp_vrf.textfsm"
        for vrf in vrf_names:
            command = "show bgp vrf {} ipv4 unicast".format(vrf)
            if isinstance(output, list):
                output += log_cmd_textfsm(connection, net_dev, command, count, txt_fsm)
            else:
                output["vrf"] = vrf
                output["status"] = "No BGP Data"
                net_dev.show_for_xls["gather_bgp"] = [output]
    else:
        output = log_cmd_textfsm(connection, net_dev, command, count)
    if isinstance(output, list):
        net_dev.show_for_xls["gather_bgp"] = output
    elif isinstance(output, str):
        output = {}
        output["status"] = "No BGP Data"
        net_dev.show_for_xls["gather_bgp"] = [output]


## XR Ready
def gather_inventory(connection, net_dev, count):
    command = "show inventory"
    txt_tmpl = None
    if net_dev.parse_method == "cisco_xr":
        txt_tmpl = "ntc-templates/templates/cisco_ios_show_inventory.textfsm"
        txt_tmpl = mod_dir_based_on_os(txt_tmpl)
    output = log_cmd_textfsm(connection, net_dev, command, count, txt_tmpl)
    if isinstance(output, list):
        net_dev.show_for_xls["gather_inventory"] = output.copy()
        update_sfp_cout(net_dev, output)
    elif isinstance(output, str):
        output = {}
        output["status"] = "Issue, have Developer check the script"
        net_dev.show_for_xls["gather_inventory"] = [output]


def update_sfp_cout(net_dev, inventory_list):
    sfp_count = 0
    for item in inventory_list:
        if "sfp" in item["descr"].lower():
            sfp_count +=1
    net_dev.sfp_count=sfp_count


## XR Ready
def gather_commands(connection, net_dev, other_shows, count=0):
    """
    Logs the other show commands requested by the user, in "Commands" sheet
    """
    for command in other_shows:
        if VERBOSE:
            print_net_dev_msg(net_dev, "Capturing '{}' as raw text".format(command))
        output = connection.send_command(command)
        net_dev.user_rqstd_show[command] = output


####Initial Setup Fucntions
def read_settings_sheet(ws_obj, start_row=5,end_row=15):
    """
    Read all the Settings from the Settings sheet and generate the appropriate Dictionary.
    """
    t_dict={}
    for i in range(start_row, end_row+1):
        prompt = rw_cell(ws_obj, i, 1).lower()
        t_dict[prompt] = rw_cell(ws_obj, i, 3)
        if isinstance(t_dict[prompt], int):
            continue
        elif t_dict[prompt] in ["Yes", "YES", "yes"]:
            t_dict[prompt] = True
        else:
            t_dict[prompt] = False
    return t_dict

def read_global_variables(ws_obj):
    t_dict = {}
    t_dict["username"] = rw_cell(ws_obj, 1, 2)
    t_dict["password"] = rw_cell(ws_obj, 2, 2)
    t_dict["secret"] = rw_cell(ws_obj, 3, 2)
    if not t_dict["secret"]:
        t_dict["secret"] = t_dict["password"]
    t_dict["output_dir"] = verify_path(rw_cell(ws_obj, 4, 2))
    t_dict["output_file"] = rw_cell(ws_obj, 5, 2)
    if t_dict["output_file"]:
        t_dict["output_file"] = add_xls_tag(t_dict["output_file"])
    else:
        print("Error:\tNo Output file value was entered.")
        print("\tPlease enter an output file name and try again.")
        sys.exit()
    return t_dict

def cell_iter_to_list(cell_iter, ignore_empty_cell):
    t_list=[]
    for cell in cell_iter:
        if not ignore_empty_cell:
            t_list.append(cell.value)
        elif cell.value:
            t_list.append(cell.value)
    return t_list

def update_with_cli_args(glbl_var, cli_arg):
    for key, val in cli_arg.items():
        if key in list(glbl_var.keys()) and val:
            if key =="output_dir":
                glbl_var[key] = verify_path(val)
            else:
                glbl_var[key] = val

def get_setup_vars(wb_obj, cli_arg):
    """
    Reads all the setup variables from the XLS document.
    """
    return_dict = {}
    t_list = []
    t_dict = {}
    # Gathers the Commands to capture
    return_dict["other_commands"] = cell_iter_to_list(wb_obj["Commands"]["A"], True)
    # Gathers the Settings, on which functions to do.
    return_dict["settings"] = read_settings_sheet(wb_obj["Settings"])
    # Adds all the global parameters
    t_glbl_dict =read_global_variables(wb_obj["Main"])
    ##Update override with any CLI Arguments
    update_with_cli_args(t_glbl_dict, cli_arg)
    return_dict["global"] = t_glbl_dict
    return return_dict


def read_network_devices(wb_obj, dflt_creds):
    """
        Takes WB and checks the Main to create NetworkDevices and return
        a list of Network_Devices
    """
    sheet_obj = get_xls_sheet(wb_obj, "Main")
    return_list = []
        # Read all the devices
    for i in range(8, sheet_obj.max_row + 1):
        host = rw_cell(sheet_obj, i, 1)
        if host:
            active = rw_cell(sheet_obj, i, 2)
            if not active:
                active = "Yes"
            parse_method = rw_cell(sheet_obj, i, 3)
            if not parse_method:
                parse_method = "autodetect"
            protocol = rw_cell(sheet_obj, i, 4)
            port_override = rw_cell(sheet_obj, i, 5)
            user_name = rw_cell(sheet_obj, i, 6)
            user_pass = rw_cell(sheet_obj, i, 7)
            # If no creds were entered then use the default global creds
            if not user_name:
                user_name = dflt_creds["username"]
            if not user_pass:
                user_pass = dflt_creds["password"]
            net_device = NetworkDevice(host,
                                       user_name,
                                       user_pass,
                                       dflt_creds["secret"],
                                       parse_method,
                                       protocol,
                                       port_override,
                                       active
                                       )
            net_device.main_col = i
            return_list.append(net_device)
    return return_list


def update_ntc_templ_path():
    """
    Adds 'NET_TEXTFSM' variable to the os.environ, required to utilize
    textfsm with Netmiko.
    """
    ntc_dir = "ntc-templates/templates"
    os.environ["NET_TEXTFSM"] = str(Path(os.getcwd())/Path(ntc_dir))


def get_other_shows(wb_obj):
    """
    Obtains the "show " commands from the WB, returns dict with Commands
    as keys and None for values.
    """
    wb_sheet = get_xls_sheet(wb_obj, "Commands")
    return_dict = {}
    for cell in wb_sheet["A"]:
        if cell.value:
            return_dict[(cell.value)] = None
    return return_dict


def get_json_data_from_file(file_name):
    """
    Reads a json file and imports all the values.
    """
    with open(file_name) as file:
        return_value = json.load(file)
    return return_value


"""###Connection and logging of the devices###"""


def connect_devices(net_devices, setup_vars):
    """
    Connects to devices and calls all the functions that log all the data.
    """
    thread_list = []
    max_threads = setup_vars["settings"]["max_threads"]
    for n, device in enumerate(net_devices):
        thread_list.append(Thread(target=con_thread, args=(device, setup_vars, n)))
        if len(thread_list) == max_threads or n == len(net_devices) - 1:
            for thread in thread_list:
                thread.start()
            for thread in thread_list:
                thread.join()
            thread_list = []


def con_thread(net_dev, setup_vars, n):
    """
    Function that contains all the functions to be Completed while
    multithreading, all the functions that gather data fromt the device.
    """
    other_shows = setup_vars["other_commands"]
    settings = setup_vars["settings"]
    if net_dev.active == "Yes":
        net_dev.collection_time = get_current_time()
        start_time = time.time()
        conn = connect_single_device(net_dev, n)
        # Add any command function captures here
        if conn != None:
            # Start CLI Log if True
            if RAW_CLI_OUTPUT:
                start_connection_log(conn, net_dev, setup_vars["global"]["output_dir"])
            # Sections are executed based on Settings
            if settings["gather_version"]:
                try:
                    gather_version(conn, net_dev, n)
                except Exception as e:
                    if VERBOSE:
                        print(60*"*"+"\n",net_dev.host, "| Issue with Gather Version", "\n"+60*"*")
                    net_dev.add_detected_error(e)
            if settings["gather_arp"]:
                try:
                    gather_arp(conn, net_dev, n)
                except Exception as e:
                    if VERBOSE:
                        print(60*"*"+"\n",net_dev.host, "| Issue with Gather ARP", "\n"+60*"*")
                    net_dev.add_detected_error(e)
            if settings["gather_mac"]:
                try:
                    gather_mac(conn, net_dev, n)
                except Exception as e:
                    if VERBOSE:
                        print(60*"*"+"\n",net_dev.host, "| Issue with Gather MAC", "\n"+60*"*")
                    net_dev.add_detected_error(e)
            if settings["gather_interface"]:
                try:
                    gather_interface(conn, net_dev, n)
                except Exception as e:
                    if VERBOSE:
                        print(60*"*"+"\n",net_dev.host, "| Issue with Gather Interface", "\n"+60*"*")
                    net_dev.add_detected_error(e)
            if settings["gather_cdp"]:
                try:
                    gather_cdp(conn, net_dev, n)
                except Exception as e:
                    if VERBOSE:
                        print(60*"*"+"\n",net_dev.host, "| Issue with Gather CDP", "\n"+60*"*")
                    net_dev.add_detected_error(e)
            if settings["gather_lldp"]:
                try:
                    gather_lldp(conn, net_dev, n)
                except Exception as e:
                    if VERBOSE:
                        print(60*"*"+"\n",net_dev.host, "| Issue with Gather LLDP", "\n"+60*"*")
                    net_dev.add_detected_error(e)
            if settings["gather_route"]:
                try:
                    gather_route(conn, net_dev, n)
                except Exception as e:
                    if VERBOSE:
                        print(60*"*"+"\n",net_dev.host, "| Issue with Gather Route", "\n"+60*"*")
                    net_dev.add_detected_error(e)
            if settings["gather_bgp"]:
                try:
                    gather_bgp(conn, net_dev, n)
                except Exception as e:
                    if VERBOSE:
                        print(60*"*"+"\n",net_dev.host, "| Issue with Gather BGP", "\n"+60*"*")
                    net_dev.add_detected_error(e)
            if settings["gather_inventory"]:
                try:
                    gather_inventory(conn, net_dev, n)
                except Exception as e:
                    if VERBOSE:
                        print(60*"*"+"\n",net_dev.host, "| Issue with Gather Inventory", "\n"+60*"*")
                    net_dev.add_detected_error(e)
            if settings["gather_commands"]:
                try:
                    gather_commands(conn, net_dev, other_shows, n)
                except Exception as e:
                    if VERBOSE:
                        print(60*"*"+"\n",net_dev.host, "| Issue with Gather Commands", "\n"+60*"*")
                    net_dev.add_detected_error(e)
            net_dev.active = "Completed"
        else:
            net_dev.active = "Error"
            if VERBOSE:
                print(22*"*", "Connection Error", 21*"*")
                print_net_dev_msg(net_dev,"Unable to establish a connection")
                print(60*"*")
        net_dev.elapsed_time = int(time.time() - start_time)


def start_connection_log(conn, net_dev, log_path):
    """Starts logging in Append Mode"""
    log_path = Path(log_path)
    log_path = (log_path/"raw_cli"/(net_dev.host+"_raw_cli.log"))
    conn.open_session_log(str(log_path), "append")
    if VERBOSE:
        print_net_dev_msg(net_dev,"Session Logging has been enabled")

def connect_single_device(net_dev, count):
    """
    Attempts to Connect to Device.
    Returns False if failed to connect otherwise it will return
    ConnectionHandler variable. Will also submit "term len 0 " command.
    """
    try:
        if VERBOSE:
            print_net_dev_msg(net_dev,"Starting Connection")
        conn = netmiko.ConnectHandler(**net_dev.connection)
        conn.enable()
        if net_dev.parse_method == "extreme_exos":
            conn.send_command("dis clip")
        else:
            conn.send_command("term len 0")
        get_hostname(conn, net_dev)
        if VERBOSE:
            print_net_dev_msg(net_dev, "Hostname is: {}".format( str(net_dev.hostname)))
        return conn
    except Exception as e:
        net_dev.conn_error_detected(e)
        return None


###Get and Parse data functions
def log_cmd_textfsm(connection, net_dev, command, count, txtfsm_tmpl=None):
    """
    Logs the Command with the textfsm option enabled, it also uses
    a template if needed. Also adds the output to the NetworkDevice
    show_output_json variable
    """
    if VERBOSE:
        print_net_dev_msg(net_dev, "Capturing '{}' with TextFSM Enabled".format(command))
    if txtfsm_tmpl:
        mod_dir_based_on_os(txtfsm_tmpl)
    output = connection.send_command(command, use_textfsm=True, textfsm_template=txtfsm_tmpl)
    if isinstance(output, list):
        net_dev.show_output_json[command] = output.copy()
    else:
        net_dev.show_output_json[command] = output
    return output


def show_proc_cpu(connection, net_dev, count):
    """
    Sends Commands "show processes cpu" and logs to variables of
    the NetworkDevice.
    """
    dev_type = net_dev.parse_method
    command = "show processes cpu"
    output = log_cmd_textfsm(connection, net_dev, command, count)
    try:
        # Need this for devices with multiple cores.
        if len(output)>1 and isinstance(output, list):
            output = join_cpu_list(output)
        if dev_type in ["cisco_ios", "cisco_nxos"]:
            net_dev.cpu_5_sec = output[0]['cpu_5_sec']
            net_dev.cpu_1_min = output[0]['cpu_1_min']
            net_dev.cpu_5_min = output[0]['cpu_5_min']
        elif dev_type == "cisco_xr":
            net_dev.cpu_1_min = output[0]['cpu_1_min']
            net_dev.cpu_5_min = output[0]['cpu_5_min']
            net_dev.cpu_15_min = output[0]['cpu_15_min']
    except Exception as e:
        net_dev.add_detected_error(e)


def join_cpu_list(cpu_list):
    """
    Joins a cpu list when more then one CPU exists.
    returns the values with a comma deliminator
    """
    return_dict = {}
    for i,line in enumerate(cpu_list):
        for key, val in line.items():
            if key in return_dict:
                return_dict[key] += ', '+'Core '+str(i+1)+': '+val+'%'
            else:
                return_dict[key] = 'Core '+str(i+1)+': '+val+'%'
    return return_dict



def get_vrf_names(net_dev, connection, count):
    """
    Gathers the VRF names from the connection. If the names were already
    gathered then it returns the list from the NetworkDevice variable of
    vrf_names.
    """
    try:
        vrf_names = net_dev.vrf_names
    except AttributeError:
        vrf_names = ["global"]
        txt_tmpl = None
        command = "show vrf"
        if net_dev.parse_method == "cisco_xr":
            vrf_names = ["default"]
            command += " all"
            txt_tmpl = "ntc-templates/test_tmpl/cisco_xr_show_vrf_all.textfsm"
            txt_tmpl = mod_dir_based_on_os(txt_tmpl)
        output = log_cmd_textfsm(connection, net_dev, command, count, txt_tmpl)

        if VERBOSE:
            print_net_dev_msg(net_dev, "Parsing VRFs")
        if isinstance(output, list):
            for vrf in output:
                if vrf['name'] not in vrf_names:
                    vrf_names.append(vrf['name'])
        elif isinstance(output, str):
            net_dev.add_error_msg("Issue with the Gather VRF Names, seems to be an issue with 'show vrf', check textfsm template., it is not parsing the data into a list, get a string.")
        net_dev.vrf_names = vrf_names
    return vrf_names


def get_xr_locations(net_dev, connection, count):
    """
    Gathers the locations
     names from the connection. If the names were already
    gathered then it returns the list from the NetworkDevice variable of
    vrf_names.
    """
    try:
        return net_dev.xr_locations
    except AttributeError:
        if VERBOSE:
            print_net_dev_msg(net_dev,"Parsing XR Device Locations")
        command = "show l2vpn forwarding bridge-domain : mac-address location ?"
        output = connection.send_command(command)
        connection.send_command("")
        net_dev.xr_locations = parse_locations_frm_prmpt(output)
        return net_dev.xr_locations


def get_xr_bg_grp_dmns(net_dev, connection, count):
    """
    Gathers the Bridge group and domain and returns them with a colon
    """
    try:
        bg_grp_dmns = net_dev.bg_grp_dmns
    except AttributeError:
        if VERBOSE:
            print_net_dev_msg(net_dev,"Parsing BGP Group and Domain")
        net_dev.bg_grp_dmns = []
        command = "show l2vpn bridge-domain"
        txt_tmpl = "ntc-templates/test_tmpl/cisco_xr_show_l2vpn_forwarding_bridge_info.textfsm"
        output = log_cmd_textfsm(connection, net_dev, command, count, txt_tmpl)
        if isinstance(output, list):
            for i in output:
                net_dev.bg_grp_dmns.append("{}:{}".format(i["bridge_group"], i["bridge_domain"]))
    return net_dev.bg_grp_dmns


def get_trunk_dict(device, connection):
    """
    Parses the 'show int trunk' response, only works on cisco_ios and cisco_nxos
    """
    trunk_all_info = connection.send_command("show int trunk").split('\n')
    vlans_native_list, vlans_allowed_list, vlans_forwarding_list = [], [], []
    vlans_not_pruned_list, vlans_err_disabled_list = [], []
    x = 0

    if device.parse_method == "cisco_ios":
        for line in trunk_all_info:
            if line not in ["", " ", "\n"]:
                if x == 0:
                    # Find the first reference to the word 'port' which will house native vlan data
                    first_word = line.split(" ")[0]
                    if first_word.lower() == "port":
                        x = x + 1
                elif x == 1:
                    first_word = line.split(" ")[0]
                    if first_word.lower() == "port":
                        x = x + 1
                    # Add lines to native_vlan_list:
                    else:
                        vlans_native_list.append(line)
                elif x == 2:
                    first_word = line.split(" ")[0]
                    # Increment counter when word 'Port' is found again
                    if first_word.lower() == "port":
                        x = x + 1
                    # Add lines to vlan_allowed_list:
                    else:
                        vlans_allowed_list.append(line)
                elif x == 3:
                    first_word = line.split(" ")[0]
                    # Increment counter when word 'Port' is found again
                    if first_word.lower() == "port":
                        x = x + 1
                    # Add lines to vlan_active_list:
                    else:
                        vlans_forwarding_list.append(line)
                elif x == 4:
                    # Add lines to vlan_active_list:
                    vlans_not_pruned_list.append(line)

    if device.parse_method == "cisco_nxos":
        for line in trunk_all_info:
            if line != "" and left(line, 1) != " " and line != "\n":
                if left(line, 3) != "---" and left(line, 7) != "Feature":
                    if x == 0:
                        # Find the first reference to the word 'port'
                        # which will house native vlan data.
                        first_word = line.split(" ")[0]
                        if first_word.lower() == "port":
                            x = x + 1
                    elif x == 1:
                        # Find second instance of the word 'port'
                        # (VLANS Allowed) and increment counter
                        first_word = line.split(" ")[0]
                        if first_word.lower() == "port":
                            x = x + 1
                        # Add lines to native_vlan_list:
                        else:
                            vlans_native_list.append(line)
                    elif x == 2:
                        # Increment counter when word 'Port'
                        # is found again (ERR Disabled)
                        first_word = line.split(" ")[0]
                        if first_word.lower() == "port":
                            x = x + 1
                        # Add lines to vlan_err_disabled_list:
                        else:
                            vlans_allowed_list.append(line)
                    elif x == 3:
                        # Increment counter when word 'Port'
                        # is found again (ERR Disabled)
                        first_word = line.split(" ")[0]
                        if first_word.lower() == "port":
                            x = x + 1
                        # Add lines to vlan_err_disabled_list:
                        else:
                            vlans_err_disabled_list.append(line)
                    elif x == 4:
                        first_word = line.split(" ")[0]
                        # Increment counter when word 'Port'
                        # is found again (STP Forwarding)
                        if first_word.lower() == "port":
                            x = x + 1
                        # Add lines to vlan_active_list:
                        else:
                            vlans_forwarding_list.append(line)
                    elif x == 5:
                        # Add lines to vlan_active_list:
                        vlans_not_pruned_list.append(line)

    return {'vlans_native': vlans_native_list,
            'vlans_allowed': vlans_allowed_list,
            'vlans_err_disabled': vlans_err_disabled_list,
            'vlans_forwarding': vlans_forwarding_list,
            'vlans_not_pruned': vlans_not_pruned_list,
            }


def get_trunk_details(if_name, trunk_dict, key_value, net_dev):
    """
    Provides the trunk details for partivular key_value:
    vlans_native, vlans_allowed, vlans_not_pruned
    """
    try:
        for x in trunk_dict[key_value]:
            interface, value = "", ""
            if key_value != "vlans_native":
                interface = x.split()[0]
                value = x.split()[1]
            else:
                if net_dev.parse_method == "cisco_ios":
                    interface = x.split()[0]
                    value = x.split()[4]
                elif net_dev.parse_method == "cisco_nxos":
                    interface = x.split()[0]
                    value = x.split()[1]
            if if_name.lower() == interface.lower():
                return value
    except Exception as e:
        net_dev.add_detected_error(e)
        if VERBOSE:
            print_net_dev_msg(net_dev, "Error getting trunk info from device: "+str(e))
        return "Error, review Errors Log"


def count_interfaces(if_dictionary):
    """
    Count Number of interfaces per device, input has to be
    "show interface" with TEXTFSM=True
    """
    return_dict = {
        "Ethernet": {"count": 0, "active": 0},
        "FastEthernet": {"count": 0, "active": 0},
        "GigabitEthernet": {"count": 0, "active": 0},
        "TenGigEthernet": {"count": 0, "active": 0},
        "TwentyFiveGigEthernet": {"count": 0, "active": 0},
        "FortyGigEthernet": {"count": 0, "active": 0},
        "HundredGigEthernet": {"count": 0, "active": 0},
        "Serial": {"count": 0, "active": 0},
        "Subinterfaces": {"count": 0, "active": 0},
        "Tunnel": {"count": 0, "active": 0},
        "Port-channel": {"count": 0, "active": 0},
        "Loopback": {"count": 0, "active": 0},
        "VLAN": {"count": 0, "active": 0}
    }
    for i in if_dictionary:
        split_if = i['interface'].split(".")
        if len(split_if) == 1:
            if left(i['interface'], 3).lower() == "eth":
                return_dict["Ethernet"]["count"] += 1
                if i['link_status'] == "up":
                    return_dict["Ethernet"]["active"] += 1
            elif left(i['interface'], 3).lower() == "fas":
                return_dict["FastEthernet"]["count"] += 1
                if i['link_status'] == "up":
                    return_dict["FastEthernet"]["active"] += 1
            elif left(i['interface'], 3).lower() == "gig":
                return_dict["GigabitEthernet"]["count"] += 1
                if i['link_status'] == "up":
                    return_dict["GigabitEthernet"]["active"] += 1
            elif left(i['interface'], 3).lower() == "ten":
                return_dict["TenGigEthernet"]["count"] += 1
                if i['link_status'] == "up":
                    return_dict["TenGigEthernet"]["active"] += 1
            elif left(i['interface'], 3).lower() == "twe":
                return_dict["TwentyFiveGigEthernet"]["count"] += 1
                if i['link_status'] == "up":
                    return_dict["TwentyFiveGigEthernet"]["active"] += 1
            elif left(i['interface'], 3).lower() == "for":
                return_dict["FortyGigEthernet"]["count"] += 1
                if i['link_status'] == "up":
                    return_dict["FortyGigEthernet"]["active"] += 1
            elif left(i['interface'], 3).lower() == "hun":
                return_dict["HundredGigEthernet"]["count"] += 1
                if i['link_status'] == "up":
                    return_dict["HundredGigEthernet"]["active"] += 1
            elif left(i['interface'], 6).lower() == "serial":
                return_dict["Serial"]["count"] += 1
                if i['link_status'] == "up":
                    return_dict["Serial"]["active"] += 1
            elif left(i['interface'], 3).lower() == "tun":
                return_dict["Tunnel"]["count"] += 1
                if i['link_status'] == "up":
                    return_dict["Tunnel"]["active"] += 1
            elif left(i['interface'], 5).lower() == "port-":
                return_dict["Port-channel"]["count"] += 1
                if i['link_status'] == "up":
                    return_dict["Port-channel"]["active"] += 1
            elif left(i['interface'], 3).lower() == "loo":
                return_dict["Loopback"]["count"] += 1
                if i['link_status'] == "up":
                    return_dict["Loopback"]["active"] += 1
            elif left(i['interface'], 3).lower() == "vla":
                return_dict["VLAN"]["count"] += 1
                if i['link_status'] == "up":
                    return_dict["VLAN"]["active"] += 1
        elif len(split_if) == 2:
            return_dict["Subinterfaces"]["count"] += 1
            if i['link_status'] == "up":
                return_dict["Subinterfaces"]["active"] += 1
    return return_dict


def get_vrf_interfaces_dict(device, conn, count):
    """
    This function will parse VRF information to get interface and VRF
    info.
    THIS IS MESSY, IT NEEDS TO BE CLEANED
    """
    command = "show vrf"
    if device.parse_method == "cisco_xr":
        vrf_names = get_vrf_names(device, conn, count)
        return_list = []
        for vrf in vrf_names:
            if vrf != "default":
                vrf_dict = {}
                command = "show vrf " + vrf + " detail"
                output = conn.send_command(command)
                start_log = False
                for line in output.split("\n"):
                    if line == "Interfaces:":
                        start_log = True
                    elif "Address" in line:
                        start_log = False
                        break
                    elif start_log:
                        vrf_dict["name"] = vrf
                        vrf_dict["interfaces"] = line[2:]
                        return_list.append(vrf_dict.copy())
        return return_list
    if device.parse_method == "cisco_nxos":
        command += " interface"

    output = log_cmd_textfsm(conn, device, command, count)
    if "Invalid input detected at" in output:
        rtr_str = "Invalid Input"
        return rtr_str
    if isinstance(output, str):
        device.add_error_msg("Issue with 'get_vrf_interfaces_dict', seems to be an issue with 'show vrf', check textfsm template., it is not parsing the data into a list, get a string. \nString Output:\n"+output+"\n")
    return output


####Save Functions
def save_device_data(net_devices, wb_obj, setup_vars, key_map):
    """
    Runs all the Save functions for the network devices
    """
    glbl_set = setup_vars["global"]
    settings = setup_vars["settings"]
    for net_dev in net_devices:
        net_dev.update_outdir_outfile(glbl_set["output_dir"])
        if net_dev.active == "Completed" and net_dev.elapsed_time:
            gather_results_to_wb(wb_obj, net_dev, key_map)
            if settings["gather_commands"]:
                save_other_shows_to_txt(net_dev)
        write_dev_vars_to_wb(wb_obj, net_dev, key_map["device_info_map"])
        save_dev_show_json_data(net_dev)
        add_err_msgs_to_wb(net_dev, wb_obj)
    save_xls(wb_obj, glbl_set["output_file"], glbl_set["output_dir"])


def save_dev_show_json_data(net_dev):
    """
    Saves the output of the "gather_" functions as json or raw depending
    on whether textfsm applied. Creates a new file for every device in
    a subdirectory of the "output directory".
    """
    json_data = net_dev.show_output_json
    json_f_name = net_dev.json_out_file
    output_dir = net_dev.out_dir + "JSON/"
    if json_data:
        output_dir = verify_path(output_dir)
        spacer = gen_spacer("-", 2)
        cmd_output_spacer = gen_spacer("#", 1)
        file_name = output_dir + json_f_name
        with open(file_name, "w+") as filehandle:
            write_str = spacer + center_string("Connected to " + net_dev.host)
            write_str += "\n" + center_string("Hostname is: " + net_dev.hostname)
            write_str += "\n" + spacer
            spacer = gen_spacer()
            filehandle.write(write_str)
            for show_cmd, output in json_data.items():
                str_output = json.dumps(output, indent=1)
                write_str = spacer + center_string("****** " + show_cmd + " ******")
                write_str += "\n" + cmd_output_spacer + str_output
                write_str += "\n" + cmd_output_spacer
                filehandle.write(write_str)
            write_str = spacer + "*" * 20 + "\tEnd of File\t" + "*" * 20
            filehandle.write(write_str)

def print_net_dev_msg(net_dev, msg):
    line1 = str(net_dev.main_col)
    if len(line1)<8:
        line1 += (8-len(line1))*" "
    line1 = " "+line1
    line2 = str(net_dev.host)
    if len(line2)<15:
        line2 += (15-len(line2))*" "
    print(line1, "|", line2, "|", msg)


def write_dev_vars_to_wb(wb_obj, device, var_loc):
    """
    Writes all the data to the "Main" worksheet
    """
    wb_sheet = wb_obj["Main"]
    dev_vars = vars(device)
    row = device.main_col
    for key, col in var_loc.items():
        if key in list(dev_vars.keys()):
            if key == "interface_count" and dev_vars[key]:
                for interf, c_key in col.items():
                    values = dev_vars[key][interf]
                    rw_cell(wb_sheet, row, c_key["count"], True, str(values["count"]))
                    rw_cell(wb_sheet, row, c_key["active"], True, str(values["active"]))
            elif isinstance(col, int):
                rw_cell(wb_sheet, row, col, True, str(dev_vars[key]))


def save_other_shows_to_txt(net_dev):
    """
    Saves the results to the "other" show commands entered on the
    spreadsheet. creates a new file for every device in a subdirectory
    of the "output directory"
    """
    if net_dev.user_rqstd_show:
        file_name = verify_path(net_dev.out_dir + "show_cmd_captures/")
        spacer = gen_spacer()
        cmd_output_spacer = gen_spacer("#", 1)
        file_name += net_dev.hostname + ".txt"
        with open(file_name, "w+") as filehandle:
            write_str = spacer
            write_str += center_string("Connected to " + net_dev.host) + "\n"
            write_str += center_string("Hostname is: " + net_dev.hostname)
            write_str += "\n" + spacer
            filehandle.write(write_str)
            for show_cmd in net_dev.user_rqstd_show.keys():
                write_str = show_cmd + "\n"
                filehandle.write(write_str)
            filehandle.write(spacer)
            for show_cmd, output in net_dev.user_rqstd_show.items():
                write_str = center_string("****** " + show_cmd + " ******")
                write_str += "\n" + cmd_output_spacer + output + "\n"
                write_str += cmd_output_spacer + spacer
                filehandle.write(write_str)
            write_str = "*" * 23 + " End of File " + "*" * 24
            filehandle.write(write_str)


def gather_results_to_wb(wb_obj, net_dev, key_map):
    """
    Writes the Simple show commands to the work book based on the
    key_map information.
    """
    headers_key = map_headers(wb_obj)
    show_results = net_dev.show_for_xls
    d_parse = net_dev.parse_method
    hostname = net_dev.hostname
    for setting, list_of_shows in show_results.items():
        if isinstance(list_of_shows, list):
            sheet = list(key_map[d_parse][setting].keys())[0]
            c_mapper = key_map[d_parse][setting][sheet]
            for value in list_of_shows:
                row = next_available_row(wb_obj[sheet])
                rw_cell(wb_obj[sheet], row, 1, True, hostname)
                for key, c_name in c_mapper.items():
                    if key in value.keys():
                        col = headers_key[sheet][c_name]
                        wr_val = value[key]
                        rw_cell(wb_obj[sheet], row, col, True, wr_val)
        else:
            print("Error", list_of_shows)


def add_err_msgs_to_wb(net_dev, wb_obj):
    """
    Adds all the Registered Errors to the WorkSheet.
    """
    sheet_obj = wb_obj["Errors"]
    err_msgs = net_dev.error_msgs
    row = next_available_row(sheet_obj)
    hostname = net_dev.hostname
    if not hostname:
        hostname = net_dev.host
    for i, msg in enumerate(err_msgs):
        row += i
        rw_cell(sheet_obj, row, 1, True, hostname)
        rw_cell(sheet_obj, row, 2, True, msg[1])
        rw_cell(sheet_obj, row, 3, True, msg[0])


def save_xls(wb_obj, file_name=None, output_dir=None):
    """
    Saves the WorkBook to provided Directory and File Name.
    If no File Name and/or Directory provided it will save to im
    """
    file_save_string = ""
    if output_dir:
        output_dir = verify_path(output_dir)
        file_save_string = output_dir
    if file_name:
        file_save_string += file_name
    else:
        file_save_string += INPUT_FILE_NAME[:-5] + "_new.xlsx"
    print("saving the file to: " + file_save_string)
    wb_obj.save(file_save_string)


###Helper Functinos###
def parse_locations_frm_prmpt(raw_str):
    """
    Parses the ? prompt to get the locations
    """
    locations_list = []
    for line in raw_str.split("\n"):
        print('"' + str(line) + '"')
        if line and ":" not in line and "WORD" not in line and "ncomplete" not in line:
            for i in line.split(" "):
                if i:
                    locations_list.append(i)
                    break
    return locations_list


def find_val_in_col(value, column):
    """
    Returns the Row value in which the "value" was found
    """
    for n, cell in enumerate(column):
        if cell.value == value:
            return n + 1
    return 0


def gen_spacer(spacer_char="-", nl=2):
    """
    Returns a spacer string with 60 of designated character, "-" is default
    It will generate two lines of 60 characters
    """
    spacer = ""
    for i in range(nl):
        spacer += spacer_char * 60
        spacer += "\n"
    return spacer


def get_current_time(str_option="dt"):
    """
    Captures the current time and returns it. Will return both date
    and time, or just one depending on the str_option provided.
    """
    now = datetime.now()
    str_option = str_option.lower()
    if str_option == "dt":
        return now.strftime("%m/%d/%Y") + ", " + now.strftime("%H:%M:%S")
    if str_option == "d":
        return now.strftime("%m/%d/%Y")
    if str_option == "t":
        return now.strftime("%H:%M:%S")
    return "Invalid selection.  Choose d for date, t for time, or dt for date + time."


def get_hostname(conn, net_dev):
    """
    Gets Hostname from the Connection and saves it to the device.
    """
    t_hm = conn.find_prompt()[:-1]
    if net_dev.parse_method == "cisco_xr":
        net_dev.hostname = t_hm.split(":")[1]
    else:
        net_dev.hostname = t_hm


def map_headers(wb_obj):
    """
    Returns a nested dictionary with the location and name of each
    column header
    """
    sheets = wb_obj.sheetnames
    return_value = {}
    ignore_this = ["Main", "Commands", "Settings", "Errors"]
    for sheet in sheets:
        if sheet not in ignore_this:
            row = wb_obj[sheet][1]
            sheet_mapper = {}
            for count, cell in enumerate(row):
                sheet_mapper[cell.value] = count + 1
            return_value[sheet] = sheet_mapper
    return return_value


def add_xls_tag(file_name):
    """
    Check the file_name to ensure it has ".xlsx" extension, if not add it
    """
    if file_name[:-5] != ".xlsx":
        return file_name + ".xlsx"
    else:
        return file_name


def next_available_row(sheet_obj, col='A'):
    """
    Returns the number of the next available Row, it determines
    avaibaility based on whether there is an entry for hostname
    """
    column = sheet_obj[col]
    for index, cell in enumerate(column):
        if cell.value is None:
            return index + 1
    return len(column) + 1


def get_xls_sheet(wb_obj, sheet_name):
    """
    Returns the Worksheet of provided Name
    """
    sheet_obj = wb_obj[sheet_name]
    sheet_obj.protection.sheet = False
    return sheet_obj


def format_uptime(uptime):
    """
    Returns the Formated uptime.
    """
    str_years, str_weeks, str_days, str_hours, str_minutes = 0, 0, 0, 0, 0
    str_input = uptime.split(",")
    for i in str_input:
        i = i.strip()
        str_split = i.split(" ")
        if left(str_split[1], 3) == "yea":
            str_years = int(str_split[0])
        if left(str_split[1], 3) == "wee":
            str_weeks = int(str_split[0])
        if left(str_split[1], 3) == "day":
            str_days = int(str_split[0])
        if left(str_split[1], 3) == "hou":
            str_hours = int(str_split[0])
        if left(str_split[1], 3) == "min":
            str_minutes = int(str_split[0])
    if str_days > 365:
        years = str_days / 365
        if not years.is_integer():
            years = int(str(years).split(".")[0])
        str_days = str_days - years * 365
        str_years = str_years + years
    if str_days > 7:
        weeks = str_days / 7
        if not weeks.is_integer():
            weeks = int(str(weeks).split(".")[0])
        str_days = str_days - weeks * 7
        str_weeks = str_weeks + weeks
    if str_weeks > 52:
        years = str_weeks / 52
        if not years.is_integer():
            years = years.split(".")
            years = years[0]
        str_weeks = str_weeks - years * 52
        str_years = str_years + years

    return (str(str_years) + "y " +
            str(str_weeks) + "w " +
            str(str_days) + "d " +
            str(str_hours) + "h " +
            str(str_minutes) + "m "
            )


def get_short_if_name(interface, device_type):
    """
    Returns short if name. for cisco_ios it returns first 2 char and the
    interface number. Everything else it returns the first 3 plus number.
    """
    number = re.compile(r"(\d.*)$")
    name = re.compile("([a-zA-Z]+)")
    number = number.search(interface).group(1)
    name = name.search(interface).group(1)
    short_name = ""
    if device_type == "cisco_ios":
        short_name = left(name, 2)
    elif device_type in ["cisco_nxos", "cisco_xr"]:
        port = left(name, 3).lower()
        if port == "eth":
            short_name = left(name, 3)
        elif port == "vla":
            short_name = left(name, 4)
        elif port == "mgm":
            short_name = left(name, 4)
        else:
            short_name = left(name, 2)
    if int(left(number, 1)) >= 0 or number is None:
        short_name = short_name + str(number)
    return short_name


def left(s, amount):
    """
    Returns the left characters of amount size
    """
    return s[:amount]


def right(s, amount):
    """
    Returns the right characters of amount size
    """
    return s[-amount:]


def mid(s, offset, amount):
    """
    Returns the middle characters starting at offset of length amount
    """
    return s[offset:offset + amount]


def open_xls(xls_input_file_name):
    """
    Returns the WorkBook of specified Name
    """
    try:
        return openpyxl.load_workbook(xls_input_file_name, data_only=True)
    except Exception as e:
        print(e)
        print("Please ensure the file exists or the correct filename was entered when utilizing the \"-i\" argument.")


def cli_args():
    """Reads the CLI options provided and returns them using the OptionParser
    Will return the Values as a dictionary"""
    parser = optparse.OptionParser()
    parser.add_option('-v','--verbose',
                      dest="verbose",
                      default=False,
                      action="store_true",
                      help="Enable Verbose Output"
                      )
    parser.add_option('-r','--raw_cli_output',
                      dest="raw_cli_output",
                      default=False,
                      action="store_true",
                      help="Capture the raw CLI output"
                      )
    parser.add_option('-i','--input_file',
                      dest="input_file",
                      default="GetInventory - Default.xlsx",
                      action="store",
                      help="Input file name of excel sheet"
                      )
    parser.add_option('-o','--output_file',
                      dest="output_file",
                      action="store",
                      help="Output file name of excel sheet"
                      )
    parser.add_option('-d','--output_directory',
                      dest="output_dir",
                      action="store",
                      help="Output Directory of excel sheet"
                      )
    parser.add_option('-u','--username',
                      dest="username",
                      action="store",
                      help="Global Username"
                      )
    parser.add_option('-p','--password',
                      dest="password",
                      action="store",
                      help="Global Password"
                      )
    parser.add_option('-s','--secret',
                      dest="secret",
                      action="store",
                      help="Global Secret"
                      )


    options, remainder = parser.parse_args()
    # Utilizing the vars() method we can return the options as a dictionary
    return vars(options)




def verify_path(output_dir):
    """
    Generates Path if it doesn't exist.
    """
    sys_type = os.name
    if sys_type == "nt" and ':' not in output_dir:
        output_dir = os.getcwd() + output_dir
    if not output_dir:
        output_dir = os.getcwd()
    if output_dir[-1] not in ["\\", "/"]:
        output_dir += "/"
    output_dir = mod_dir_based_on_os(output_dir)
    try:
        Path(output_dir).mkdir(parents=True, exist_ok=True)
    except Exception as e:
        if "Read-only" in str(e):
            print("Issue with path: " + output_dir)
            print("Error was raised: " + str(e))
            print("Suggestion: Try utilizing the full path.")
            if sys_type != "nt":
                print("If trying to use relative path please ensure that '/' is removed from leading directory name")
            print("Exiting now, please try again.")
            sys.exit()
        else:
            print("ERROR | ", e)
            sys.exit()
    return output_dir


def rw_cell(sheet_obj, row, column, write=False, value=""):
    """
    Either writes or reads to/from a cell.
    """
    if write:
        sheet_obj.cell(row=row, column=column).value = value
        return None
    return sheet_obj.cell(row=row, column=column).value


def center_string(input_str, line_length=60):
    """
    Adds space in front of the string to center it.
    """
    rtr_str = ""
    extra_space = int((line_length - len(input_str)) / 2)
    if extra_space > 0:
        rtr_str = " " * extra_space + input_str
    return rtr_str


"""########## New Network Device Class ###########"""


class NetworkDevice():
    """
    NetworkDevice class to handle and retain all output information.
    """
    supported_devices = ["cisco_xr", "cisco_ios", "cisco_nxos", "extreme_exos"]
    def __init__(
            self,
            t_host,
            t_user,
            t_pass,
            t_secret,
            t_device_type,
            t_protocol,
            port_override,
            t_active
            # ,
            # raw_cli,
            # out_dir
        ):
        self.active = t_active
        self.host = t_host
        self.hostname = ""
        self.protocol = ""
        self.parse_method = ""
        self.collection_time = ""
        self.elapsed_time = 0
        self.interface_count = {}
        self.show_output = {}
        self.user_rqstd_show = {}
        self.show_output_json = {}
        self.show_for_xls = {}
        self.cpu_1_min = ""
        self.cpu_5_min = ""
        self.cpu_5_sec = ""
        self.cpu_15_min = ""
        self.model = ""
        self.serial_number = ""
        self.uptime = ""
        self.version = ""
        self.conn_error = ""
        self.running_image = ""
        self.main_col = 0
        self.json_out_file = ""
        self.cmd_out_file = ""
        self.sfp_count = ""
        self.port = 0
        self.connection = {'device_type': 'autodetect',
                           'ip': t_host,
                           'username': t_user,
                           'password': t_pass,
                           'secret': t_secret}

        # self.log_raw_cli = raw_cli
        # self.out_dir = out_dir

        self.show_output = {}
        self.comments = []
        self.error_msgs = []
        if port_override:
            self.connection["port"] = port_override
            self.port = port_override
        # Check connection type
        if t_device_type == "autodetect":
            if self.active == "Yes":
                self.detect_device_type()
        else:
            self.connection["device_type"] = str(t_device_type)
            self.parse_method = t_device_type
        # Check if Device is supported
        if self.parse_method not in self.supported_devices:
            msg = "Unable to detect the device type, Device type detected as: "+str(self.parse_method)
            self.add_error_msg(msg)
            self.active = "Error"
        else:
            if t_protocol == "telnet":
                self.connection["device_type"] += "_telnet"
                if not self.port:
                    self.port = "23"
            else:
                self.connection["device_type"] += "_ssh"
                self.protocol = "ssh"
                if not self.port:
                    self.port = "22"
            self.connection["global_delay_factor"] = 2

    def start_connection_log(self):
        """Starts logging in Append Mode"""
        log_path = (self.out_dir_path/"raw_logs"/(self.host+"_raw_cli.log"))
        self.__add_time_stamp_to_file(log_path)
        self.connection.open_session_log(str(log_path), "append")
        if VERBOSE:
            print_net_dev_msg(self, "Session Logging has been enabled")

    def detect_device_type(self):
        """
        Detects the type of Device type.
        """
        try:
            print(self.main_col,"|",self.host,"| Detecting Device Type")
            guesser = netmiko.SSHDetect(**self.connection)
            best_match = guesser.autodetect()
            if not best_match:
                msg = "Unable to detect the device type, Device type detected as: "+str(self.parse_method)
                self.add_error_msg(msg)
                self.active = "Error"
                best_match = "Unknown"
            else:
                print(self.main_col, "|",self.host, "| Device was detected as:", best_match)
            self.connection['device_type'] = best_match
            self.parse_method = best_match
        except Exception as e:
            print(e)
            self.parse_method = "Unknown"

    def conn_error_detected(self, err_msg):
        """
        Add comment of connection error
        """
        self.conn_error = err_msg
        self.add_detected_error(err_msg)
        self.active = "Error"

    def is_socket_open(self, port):
        """
        Checks if given port is open
        """
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        try:
            sock.connect((self.host, int(port)))
            sock.shutdown(2)
            return True
        except:
            return False

    def probe_port(self, port):
        """
        Test given port, if port does not work it will test ssh and telnet
        """
        if self.is_socket_open(port):
            self.port = port
            return
        elif self.is_socket_open(22):
            return "ssh"
        elif self.is_socket_open(23):
            return "telnet"
        else:
            self.add_error_msg("None of the Ports were open. Skipping this Device.")
            self.active = "Error"
            return

    def add_comment(self, t_comment):
        """
            Adds commments that will be added to sheet
        """
        comment = str(len(t_comment) + 1) + " | "
        comment += str(t_comment)
        self.comments.append(comment)

    def add_error_msg(self, t_err_msg):
        """
            Adds commments that will be added to sheet
        """
        comment = str(len(self.error_msgs) + 1) + " | "
        comment += str(t_err_msg)
        t_time = get_current_time()
        self.error_msgs.append([comment, t_time])



    ##The following functions add more variables to the output
    ##Or handle the non simple show commands
    def read_vers_info(self):
        """
        Reads the "show version" information and updates the NetworkDevices
        Variables accordingly.
        """
        output = self.show_output_json["show version"][0]
        trantab = str.maketrans("", "", "\'\"{}[]")
        self.model = str(output["hardware"]).translate(trantab)
        self.uptime = output["uptime"]
        self.version = output["version"]
        if self.parse_method != "cisco_xr":
            self.serial_number = str(output["serial"]).translate(trantab)
            self.running_image = output["running_image"]

    def update_outdir_outfile(self, out_dir):
        """
        Updates the name of output directory and the file names for
        output json and requested commands.
        """
        if self.hostname:
            self.json_out_file = self.hostname + "_" + self.host + "_JSON_cmds.json"
            self.cmd_out_file = self.hostname + "_" + self.host + "_requested_cmds.txt"
        else:
            self.json_out_file = self.host + "_JSON_cmds.json"
            self.cmd_out_file = self.host + "_requested_cmds.txt"
        self.out_dir = out_dir

    def add_detected_error(self, e):
        exc_tb = sys.exc_info()[2]
        exc_type = sys.exc_info()[0]
        exc_line = exc_tb.tb_lineno
        full_error = traceback.format_exc()
        f_name = traceback.extract_tb(exc_tb, 1)[0][2]
        t_err_msg = "{} | Exception Type: {} | At Function: {} | Line No: {} | Error Message: {}\n{}"
        t_err_msg = t_err_msg.format(self.host, exc_type, f_name, exc_line, e, full_error)
        self.add_error_msg(t_err_msg)


###For testing Only
def testing_connection(net_devices, wb_obj, key_map):
    """
    This is only for Testing the new functinos without multithreading.
    """
    for n, device in enumerate(net_devices):
        if device.active == "Yes":
            device.collection_time = get_current_time()
            start_time = time.time()
            connection = connect_single_device(device, n)
            # Add any command function captures here
            if connection != None:
                gather_interface(connection, device, n)
                device.interface_count = count_interfaces(device.show_output_json["show interface"])
                device.active = "Completed"
                connection.disconnect()
            else:
                device.active = "Error"
            device.elapsed_time = time.time() - start_time


"""NEED TO ASK HOW WE WILL HANDLE THIS?"""


def remove_passwords(wb_obj):
    """
    Removes the passwords from the xls workbook that will be created.
    Will not remove the password from the original input xls workbook.
    """
    rw_cell(wb_obj["Main"], 1, 2, True, "")
    rw_cell(wb_obj["Main"], 2, 2, True, "")
    rw_cell(wb_obj["Main"], 3, 2, True, "")
    for n, val in enumerate(wb_obj["Main"]["F"]):
        if n > 7: # Starts at 8
            rw_cell(wb_obj["Main"], n, 6, True, "")
            rw_cell(wb_obj["Main"], n, 7, True, "")


def mod_dir_based_on_os(dir_name):
    """
    Based on type of system, it will change the '\\' to a '/', vice versa.
    if nt, then it assumes it is a windows.
    """
    if os.name == "nt":
        return dir_name.replace('/', "\\")
    return dir_name.replace('\\', "/")


if __name__ == "__main__":
    main()
